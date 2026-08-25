"""
report_core.py
--------------
Logica de leitura da planilha e montagem do contexto do relatorio semanal.
Extraida do generate_report.py original (CLI) para ser reaproveitada tanto
pela CLI quanto pelo app Streamlit, sem duplicar regras de negocio.

Nenhuma funcao aqui grava arquivo em disco nem depende de argparse - tudo
recebe caminho de arquivo (ou bytes) e devolve dados/strings.
"""

import datetime as dt
import unicodedata
from pathlib import Path

import openpyxl
from jinja2 import Environment, FileSystemLoader

TEMPLATE_DIR = Path(__file__).parent / "templates"
TEMPLATE_NAME = "email_template.html.j2"

# Abas que tem funcao fixa e por isso nao devem aparecer como opcao de
# "aba de dados da semana" no seletor da interface.
RESERVED_SHEET_NAMES = {"config", "kpis"}

BG_ODD = "#ffffff"
BG_EVEN = "#f8fdf9"

KPI_COLOR_MAP = {
    "1a serie": "#16a34a",
    "1ª serie": "#16a34a",
    "2a serie": "#2563eb",
    "2ª serie": "#2563eb",
    "3a serie": "#9333ea",
    "3ª serie": "#9333ea",
    "eja": "#ea580c",
    "mais formacao": "#0d9488",
    "plataforma": "#64748b",
}
KPI_COLOR_FALLBACK = ["#0891b2", "#ca8a04", "#dc2626", "#7c3aed", "#059669"]

STATUS_MAP = {
    "concluido": ("status-done", "Concluído"),
    "agendado": ("status-pending", "Agendado"),
    "em andamento": ("status-ongoing", "Em andamento"),
}

TAG_RULES = [
    (["cadastro"], "tag-plataforma", None),
    (["inclusao de conteudo", "inclusão de conteúdo"], "tag-plataforma", None),
    (["redacao", "redação", "enem"], "tag-redacao", None),
    (["quiz"], "tag-quiz", None),
]
DEFAULT_TAG_STYLE = ' style="background:#d1fae5;color:#065f46;font-size:10px;"'


def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )


def list_data_sheets(xlsx_path_or_buffer):
    """Retorna a lista de nomes de aba do arquivo, exceto Config/KPIs
    (que tem uso reservado e nao devem ser escolhidas como aba de dados)."""
    wb = openpyxl.load_workbook(xlsx_path_or_buffer, data_only=True, read_only=True)
    names = [
        name for name in wb.sheetnames
        if strip_accents(name).strip().lower() not in RESERVED_SHEET_NAMES
    ]
    wb.close()
    return names


def tag_for_componente(componente: str):
    normalized = strip_accents(componente).lower()
    for keywords, css_class, _ in TAG_RULES:
        for kw in keywords:
            if strip_accents(kw).lower() in normalized:
                return css_class, None
    return "tag", DEFAULT_TAG_STYLE


def status_for(value: str):
    if value is None:
        return "status-default", "-"
    normalized = strip_accents(str(value)).strip().lower()
    if normalized in STATUS_MAP:
        return STATUS_MAP[normalized]
    return "status-default", str(value).strip()


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        for cell in row:
            if cell.value and "componente curricular" in strip_accents(str(cell.value)).lower():
                return cell.row
    raise ValueError(
        "Não encontrei a linha de cabeçalho (esperava uma célula com "
        "'Componente curricular') nesta aba."
    )


def load_rows(xlsx_path_or_buffer, sheet_name: str | None):
    wb = openpyxl.load_workbook(xlsx_path_or_buffer, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = find_header_row(ws)

    raw_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ambiente, componente, feito, data, status = (list(row) + [None] * 5)[:5]
        if not any([ambiente, componente, feito, data, status]):
            continue
        raw_rows.append(
            {
                "ambiente": str(ambiente).strip() if ambiente else None,
                "componente": str(componente).strip() if componente else "",
                "feito": str(feito).strip() if feito else "",
                "data": parse_date(data),
                "status_raw": status,
            }
        )
    return raw_rows


def group_rows(raw_rows):
    groups = []
    current_name = None
    for r in raw_rows:
        if r["ambiente"]:
            current_name = r["ambiente"]
        if not groups or groups[-1]["name"] != current_name or r["ambiente"]:
            if r["ambiente"] and groups and groups[-1]["name"] == current_name:
                pass
            elif not groups or groups[-1]["name"] != current_name:
                groups.append({"name": current_name or "(sem seção)", "rows": []})
        groups[-1]["rows"].append(r)
    return groups


def build_context(groups):
    total_rows = sum(len(g["rows"]) for g in groups)
    global_index = 0
    total_concluido = 0
    total_agendado = 0
    total_outros = 0
    all_dates_concluido = []
    all_dates = []

    out_groups = []
    for g in groups:
        out_rows = []
        for r in g["rows"]:
            is_last = global_index == total_rows - 1
            bg = BG_ODD if global_index % 2 == 0 else BG_EVEN
            border = "none" if is_last else "1px solid #f0f7f1"

            tag_class, tag_style = tag_for_componente(r["componente"])
            status_class, status_label = status_for(r["status_raw"])

            if status_label == "Concluído":
                total_concluido += 1
                if r["data"]:
                    all_dates_concluido.append(r["data"])
            elif status_label == "Agendado":
                total_agendado += 1
            else:
                total_outros += 1

            if r["data"]:
                all_dates.append(r["data"])

            out_rows.append(
                {
                    "componente": r["componente"],
                    "feito": r["feito"],
                    "data_disp": r["data"].strftime("%d/%m") if r["data"] else "-",
                    "status": status_label,
                    "status_class": status_class,
                    "tag_class": tag_class,
                    "tag_style": tag_style or "",
                    "bg": bg,
                    "border": border,
                }
            )
            global_index += 1
        out_groups.append({"name": g["name"], "rows": out_rows})

    return {
        "groups": out_groups,
        "total_concluido": total_concluido,
        "total_agendado": total_agendado,
        "total_outros": total_outros,
        "dates_concluido": all_dates_concluido,
        "dates_all": all_dates,
    }


def read_config_sheet(xlsx_path_or_buffer):
    wb = openpyxl.load_workbook(xlsx_path_or_buffer, data_only=True)
    if "Config" not in wb.sheetnames:
        return {}
    ws = wb["Config"]
    config = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = strip_accents(str(row[0])).strip().lower().replace(" ", "")
        value = row[1] if len(row) > 1 else None
        config[key] = value
    return config


def format_number(value):
    if value is None or value == "":
        return "—"
    try:
        return f"{int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return str(value)


def read_kpis(xlsx_path_or_buffer):
    wb = openpyxl.load_workbook(xlsx_path_or_buffer, data_only=True)
    if "KPIs" not in wb.sheetnames:
        return []
    ws = wb["KPIs"]

    raw = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        segmento, estudantes, acessos = (list(row) + [None] * 3)[:3]
        if segmento is None or strip_accents(str(segmento)).strip().lower() == "segmento":
            continue
        if estudantes is None and acessos is None:
            continue
        raw.append((str(segmento).strip(), estudantes, acessos))

    fallback_iter = iter(KPI_COLOR_FALLBACK)
    kpis = []
    total = len(raw)
    for i, (segmento, estudantes, acessos) in enumerate(raw):
        key = strip_accents(segmento).strip().lower()
        color = KPI_COLOR_MAP.get(key) or next(fallback_iter, "#64748b")
        kpis.append(
            {
                "segmento": segmento,
                "estudantes": format_number(estudantes),
                "acessos": format_number(acessos),
                "color": color,
                "bg": BG_ODD if i % 2 == 0 else BG_EVEN,
                "border": "none" if i == total - 1 else "1px solid #e9f7ec",
            }
        )
    return kpis


def default_highlight(ctx):
    if ctx["total_concluido"] or ctx["total_agendado"]:
        return (
            f"Foram registradas <strong>{ctx['total_concluido']} ações concluídas</strong> "
            "nesta semana na plataforma."
        )
    return "Sem novidades registradas nesta semana."


def default_pending_highlight(ctx):
    if ctx["total_agendado"]:
        return f"Há <strong>{ctx['total_agendado']} ações agendadas</strong> para os próximos dias."
    return "Nenhuma ação agendada para os próximos dias."


def resolve_period(config, ctx):
    start = config.get("periodoinicio")
    end = config.get("periodofim")
    if start and end:
        return str(start), str(end)
    dates = ctx["dates_concluido"] or ctx["dates_all"]
    if dates:
        return min(dates).strftime("%d/%m/%Y"), max(dates).strftime("%d/%m/%Y")
    today = dt.date.today()
    return today.strftime("%d/%m/%Y"), today.strftime("%d/%m/%Y")


def resolve_highlights(config, ctx):
    h1 = config.get("destaque1") or default_highlight(ctx)
    h2 = config.get("destaque2") or default_pending_highlight(ctx)
    return h1, h2


def generate_report_html(xlsx_path_or_buffer, sheet_name: str):
    """Funcao principal: recebe o arquivo (caminho ou buffer, reaberto por
    chamada pois openpyxl consome o stream) e a aba de dados escolhida,
    devolve (html, resumo_dict) prontos para exibir/baixar."""
    raw_rows = load_rows(xlsx_path_or_buffer, sheet_name)
    if not raw_rows:
        raise ValueError(
            f"Nenhuma linha de dados encontrada na aba '{sheet_name}'."
        )

    groups = group_rows(raw_rows)
    ctx = build_context(groups)

    config = read_config_sheet(xlsx_path_or_buffer)
    period_start, period_end = resolve_period(config, ctx)
    highlight1, highlight2 = resolve_highlights(config, ctx)
    kpis = read_kpis(xlsx_path_or_buffer)

    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)), autoescape=False)
    template = env.get_template(TEMPLATE_NAME)

    html = template.render(
        period_start=period_start,
        period_end=period_end,
        highlight1=highlight1,
        highlight2=highlight2,
        groups=ctx["groups"],
        total_concluido=ctx["total_concluido"],
        total_agendado=ctx["total_agendado"],
        total_outros=ctx["total_outros"],
        kpis=kpis,
    )

    resumo = {
        "period_start": period_start,
        "period_end": period_end,
        "total_concluido": ctx["total_concluido"],
        "total_agendado": ctx["total_agendado"],
        "total_outros": ctx["total_outros"],
        "num_secoes": len(ctx["groups"]),
        "num_kpis": len(kpis),
        "config_encontrada": bool(config),
    }
    return html, resumo
